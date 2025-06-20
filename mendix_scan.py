import requests
import time
import argparse
import json
import re
from datetime import datetime
from http.cookies import SimpleCookie
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def initialize_session(url, provided_cookie=None, proxies=None):
    headers = {
        "Content-Type": "application/json"
    }

    payload = {
        "action": "get_session_data",
        "params": {}
    }

    try:
        if provided_cookie and "=" not in provided_cookie:
            provided_cookie = f"XASSESSIONID={provided_cookie}"

        if provided_cookie:
            headers["Cookie"] = provided_cookie

        response = requests.post(url, headers=headers, json=payload, verify=False, proxies=proxies)

        if response.status_code == 401:
            print("❌ Application requires authentication (HTTP 401).")
            return None, None, None

        if response.status_code != 200:
            print(f"❌ Failed to get session data: HTTP {response.status_code}")
            return None, None, None

        raw_cookie_headers = response.raw.headers.getlist("Set-Cookie")
        response_cookies = SimpleCookie()
        for raw_cookie in raw_cookie_headers:
            response_cookies.load(raw_cookie)

        combined_cookies = {}
        if provided_cookie:
            provided_jar = SimpleCookie()
            provided_jar.load(provided_cookie)
            for key, morsel in provided_jar.items():
                combined_cookies[key] = morsel.value

        for key, morsel in response_cookies.items():
            combined_cookies[key] = morsel.value

        cookie_header = "; ".join(f"{key}={value}" for key, value in combined_cookies.items())
        print(f"🪪 Using cookie header: {cookie_header}")

        data = response.json()
        csrf_token = data.get("csrftoken")
        if not csrf_token:
            print("❌ CSRF token not found in session data response.")
            return None, None, None

        metadata = data.get("metadata", [])
        entity_names = [entry["objectType"] for entry in metadata if "objectType" in entry]

        return cookie_header, csrf_token, entity_names

    except Exception as e:
        print(f"❌ Failed to initialize session: {e}")
        return None, None, None


def print_microflow_access_info(url, cookie, csrf_token, proxies=None):
    headers = {
        "Content-Type": "application/json",
        "Cookie": cookie,
        "X-Csrf-Token": csrf_token
    }

    payload = {
        "action": "get_session_data",
        "params": {}
    }

    response = requests.post(url, headers=headers, json=payload, verify=False, proxies=proxies)

    if response.status_code != 200:
        print(f"❌ Microflow check failed with status {response.status_code}")
        return

    try:
        data = response.json()
        microflows = data.get("microflows", {})

        if not microflows:
            print("⚠️ No microflows found.\n")
            return

        print("📦 Microflows available to this session:\n")

        for mf_def, uuid_string in microflows.items():
            try:
                parsed = json.loads(mf_def)
                entities = parsed.get("p", [])
                associations = parsed.get("a", [])
            except Exception:
                entities = []
                associations = []

            entities_str = ", ".join(entities) if entities else "None"
            associations_str = ", ".join(" → ".join(a) for a in associations) if associations else "None"

            uuid_list = [u.strip() for u in uuid_string.split(",") if u.strip()]
            flow_count = len(uuid_list)

            print(f"✅ Microflow Targets: {entities_str:<60} ➜ {flow_count} flow(s)")
            if associations:
                print(f"   🔗 Associations: {associations_str}")
            for uuid in uuid_list:
                print(f"   🆔 ID: {uuid}")

        print("")

    except Exception as e:
        print(f"❌ Failed to parse microflow data: {e}")


def extract_all_attributes_with_flags(obj):
    result = {"_guid": (obj.get("guid", ""), True)}
    for key, val in obj.get("attributes", {}).items():
        if isinstance(val, dict) and "value" in val:
            result[key] = (val["value"], val.get("readonly", False))
    return result


def retrieve_entity_data(url, cookie, csrf_token, entity_names, limit, proxies=None):
    headers = {
        "Content-Type": "application/json",
        "Cookie": cookie,
        "X-Csrf-Token": csrf_token
    }

    summary = []
    entity_data = {}

    print("🔍 Checking object counts and extracting attributes:\n")

    for entity in entity_names:
        xpath_payload = {
            "action": "retrieve_by_xpath",
            "params": {
                "xpath": f"//{entity}",
                "schema": {},
                "count": False
            }
        }

        try:
            xpath_response = requests.post(url, headers=headers, json=xpath_payload, verify=False, proxies=proxies)

            if xpath_response.status_code == 200:
                data = xpath_response.json()
                objects = data.get("objects", [])
                obj_count = len(objects)

                extracted_data = [extract_all_attributes_with_flags(obj) for obj in objects]
                total_values = sum(
                    1 for obj_data in extracted_data
                    for _, (_, readonly) in obj_data.items() if not readonly
                )

                summary.append((entity, obj_count, total_values))
                entity_data[entity] = extracted_data

                print(f"✅ {entity:<50} ➜ {obj_count} objects | {total_values} value fields (non-readonly)")

            else:
                print(f"❌ {entity:<50} HTTP {xpath_response.status_code}")
                summary.append((entity, "Error", "Error"))

        except Exception as e:
            print(f"❌ {entity:<50} Error: {e}")
            summary.append((entity, "Error", str(e)))

        time.sleep(0.2)

    return summary, entity_data


def sanitize_sheet_name(name):
    name = re.sub(r'[\[\]\*\/\\\?\:]', '', name)
    return name[:31]


def write_to_excel(summary, entity_data, output_path="SecurityScan_Report.xlsx", limit=100):
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    headers = ["Entity", "Object Count", "Non-readonly Value Fields"]
    summary_ws.append(headers)
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)

    for row in summary:
        summary_ws.append(row)

    for entity, obj_count, _ in summary:
        if obj_count == "Error":
            continue

        records = entity_data.get(entity, [])
        sheet_title = sanitize_sheet_name(entity)
        ws = wb.create_sheet(title=sheet_title)

        if not records:
            ws.append(["No data available"])
            continue

        all_keys = ["_guid"] + sorted(set().union(*(r.keys() for r in records)) - {"_guid"})
        ws.append(all_keys)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        column_nonreadonly_flags = {key: False for key in all_keys}

        for obj in records[:limit]:
            row = []
            for key in all_keys:
                value, readonly = obj.get(key, ("", True))

                if isinstance(value, list):
                    value = ", ".join(map(str, value))
                elif isinstance(value, dict):
                    value = str(value)

                if not readonly:
                    column_nonreadonly_flags[key] = True

                row.append((value, readonly))

            ws.append([val for val, _ in row])

            for idx, (_, readonly) in enumerate(row, 1):
                if not readonly:
                    ws.cell(row=ws.max_row, column=idx).font = Font(bold=True)

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for col_idx, key in enumerate(all_keys, 1):
            if column_nonreadonly_flags.get(key):
                ws.cell(row=1, column=col_idx).fill = red_fill

        for col_idx in range(1, len(all_keys) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20

        if len(records) > limit:
            ws.append([])
            ws.append([f"... Only first {limit} of {len(records)} objects shown."])

    wb.save(output_path)
    print(f"\n📄 Excel report with entity tabs saved to: {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Analyze Mendix XAS entity data and generate Excel report")
    parser.add_argument("-u", "--url", required=True, help="Mendix app URL (with or without /xas/)")
    parser.add_argument("-c", "--cookie", required=False, help="Session cookie XASSESSIONID")
    parser.add_argument("-o", "--output", required=False, help="Path to output Excel file (default: with timestamp)")
    parser.add_argument("-m", "--microflow", action="store_true", help="Include microflow access info")
    parser.add_argument("-l", "--limit", type=int, default=100, help="Max objects to retrieve per entity (default: 100)")
    parser.add_argument("-p", "--proxy", help="Optional proxy URL (e.g. http://127.0.0.1:8080)")

    args = parser.parse_args()

    # Normalize URL
    normalized_url = args.url.rstrip("/")
    if not normalized_url.endswith("/xas"):
        normalized_url = f"{normalized_url}/xas/"
    else:
        normalized_url += "/"

    # Set proxy config
    proxies = {"http": args.proxy, "https": args.proxy} if args.proxy else None

    # Set output filename
    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        output_path = f"SecurityScan_Report_{timestamp}.xlsx"

    cookie, csrf_token, entity_names = initialize_session(normalized_url, args.cookie, proxies=proxies)
    if not cookie or not csrf_token:
        print("❌ Cannot continue without valid session.")
        exit(1)

    if args.microflow:
        print_microflow_access_info(normalized_url, cookie, csrf_token, proxies=proxies)

    if entity_names:
        summary, entity_data = retrieve_entity_data(normalized_url, cookie, csrf_token, entity_names, args.limit, proxies=proxies)
        write_to_excel(summary, entity_data, output_path, args.limit)
